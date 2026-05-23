import io
import zipfile
import xml.etree.ElementTree as ET
from typing import BinaryIO
import openpyxl
from .base import BaseConverter, ConversionResult

_SUPPORTED_IMAGE_FORMATS = frozenset({"png", "jpeg", "jpg", "gif", "bmp", "svg"})

# Namespaces used in drawing XML
_NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_NS_A   = "http://schemas.openxmlformats.org/drawingml/2006/main"
_NS_R   = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"

_UNSUPPORTED_IMAGE_FORMATS = frozenset({"wmf", "emf"})


class XlsxConverter(BaseConverter):
    def accepts(self, extension: str) -> bool:
        return extension.lower() == ".xlsx"

    def convert(self, stream: BinaryIO) -> ConversionResult:
        # Buffer the entire stream so we can re-open it as a ZIP later without
        # disturbing openpyxl's read or the caller's stream position.
        stream.seek(0)
        zip_bytes = stream.read()
        stream.seek(0)

        # read_only=False so we can access merged_cells info
        wb = openpyxl.load_workbook(io.BytesIO(zip_bytes), read_only=False, data_only=True)
        parts = []
        assets = {}

        for sheet_index, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            _fill_merged_cells(ws)
            rows = list(ws.iter_rows(values_only=True))

            trimmed = _trim_cols(rows)

            insertions = []
            for anchor_row, asset_key, img_bytes in _extract_images(zip_bytes, sheet_index, sheet_name, trimmed):
                assets[asset_key] = img_bytes
                alt = asset_key.split("/")[-1]
                insertions.append((anchor_row, f"![{alt}]({asset_key})"))
            for anchor_row, comment in _find_unsupported_images(zip_bytes, sheet_index):
                insertions.append((anchor_row, comment))

            if _is_canvas(trimmed):
                for anchor_row, text in _extract_shape_texts(zip_bytes, sheet_index):
                    insertions.append((anchor_row, text))
                shape_section = _interleave([], insertions)
                cell_section = _canvas_cells_to_text(rows)
                section = "\n\n".join(s for s in [shape_section, cell_section] if s)
            else:
                section = _interleave(trimmed, insertions)

            if section:
                parts.append(f"## {sheet_name}")
                parts.append(section)

        wb.close()
        return ConversionResult(markdown="\n\n".join(parts), assets=assets)


def _find_unsupported_images(
    zip_bytes: bytes, sheet_index: int
) -> list[tuple[int, str]]:
    """Scan the raw XLSX ZIP for WMF/EMF images that openpyxl silently drops.

    openpyxl's reader discards WMF images before they reach ws._images, so we
    must inspect the ZIP's drawing relationship files directly.  We return a
    list of (anchor_row, comment_string) so the caller can place the comment
    near the right row in the output.

    anchor_row is 0-based; if the anchor position cannot be determined it
    defaults to 0 so the comment is still emitted.
    """
    results: list[tuple[int, str]] = []

    try:
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except zipfile.BadZipFile:
        return results

    with zf:
        # Step 1: find the drawing rel for this sheet.
        sheet_rels_path = f"xl/worksheets/_rels/sheet{sheet_index}.xml.rels"
        if sheet_rels_path not in zf.namelist():
            return results

        rels_xml = zf.read(sheet_rels_path)
        rels_root = ET.fromstring(rels_xml)

        drawing_target: str | None = None
        for rel in rels_root:
            # Relationship element — no namespace on element name in package rels
            rtype = rel.get("Type", "")
            if rtype.endswith("/drawing"):
                drawing_target = rel.get("Target", "")
                break

        if not drawing_target:
            return results

        # Target may be absolute ("/xl/drawings/drawing1.xml") or relative
        # ("../drawings/drawing1.xml" from xl/worksheets/).
        if drawing_target.startswith("/"):
            drawing_zip_path = drawing_target.lstrip("/")
        else:
            drawing_zip_path = "xl/" + drawing_target.lstrip("../")
        drawing_name = drawing_zip_path.split("/")[-1]
        drawing_dir = "/".join(drawing_zip_path.split("/")[:-1])

        if drawing_zip_path not in zf.namelist():
            return results

        # Step 2: read drawing rels to build embed_id → media extension mapping.
        drawing_rels_path = f"{drawing_dir}/_rels/{drawing_name}.rels"
        if drawing_rels_path not in zf.namelist():
            return results

        drawing_rels_xml = zf.read(drawing_rels_path)
        drawing_rels_root = ET.fromstring(drawing_rels_xml)

        # Map rId → file extension (lower-cased) for unsupported formats only
        unsupported_rids: dict[str, str] = {}
        for rel in drawing_rels_root:
            target = rel.get("Target", "")
            ext = target.rsplit(".", 1)[-1].lower() if "." in target else ""
            if ext in _UNSUPPORTED_IMAGE_FORMATS:
                rid = rel.get("Id", "")
                unsupported_rids[rid] = ext

        if not unsupported_rids:
            return results

        # Step 3: parse drawing XML to find which anchors use unsupported rIds.
        drawing_xml = zf.read(drawing_zip_path)
        drawing_root = ET.fromstring(drawing_xml)

        anchor_tags = (
            f"{{{_NS_XDR}}}twoCellAnchor",
            f"{{{_NS_XDR}}}oneCellAnchor",
            f"{{{_NS_XDR}}}absoluteAnchor",
        )

        for anchor in drawing_root:
            if anchor.tag not in anchor_tags:
                continue

            # Extract the embed rId from <xdr:blipFill><a:blip r:embed="rId1"/>
            blip = anchor.find(
                f".//{{{_NS_XDR}}}blipFill/{{{_NS_A}}}blip"
            )
            if blip is None:
                # Also check picture/blipFill path used by some generators
                blip = anchor.find(
                    f".//{{{_NS_A}}}blip"
                )
            if blip is None:
                continue

            embed_id = blip.get(f"{{{_NS_R}}}embed", "")
            if embed_id not in unsupported_rids:
                continue

            ext = unsupported_rids[embed_id]

            # Try to find the anchor's top-left row (from element).
            # For twoCellAnchor: <xdr:from><xdr:row>0</xdr:row>...
            # Row values in drawing XML are 0-based.
            anchor_row = 0
            from_el = anchor.find(f"{{{_NS_XDR}}}from")
            if from_el is not None:
                row_el = from_el.find(f"{{{_NS_XDR}}}row")
                if row_el is not None and row_el.text is not None:
                    try:
                        anchor_row = int(row_el.text)  # already 0-based
                    except ValueError:
                        pass

            results.append((anchor_row, f"<!-- image skipped: unsupported format ({ext}) -->"))

    return results


def _extract_images(
    zip_bytes: bytes, sheet_index: int, sheet_name: str, rows: list
) -> list[tuple[int, str, bytes]]:
    """Return (anchor_row_0based, asset_key, image_bytes) for each supported embedded image.

    Reads directly from the XLSX ZIP so that images nested inside group shapes
    (grpSp) are found — openpyxl's ws._images only surfaces direct-child pics
    and misses anything inside groups, which is common in real-world files.
    """
    results: list[tuple[int, str, bytes]] = []
    safe_name = sheet_name.replace("/", "_").replace("\\", "_")
    counter = 1

    try:
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except zipfile.BadZipFile:
        return results

    with zf:
        sheet_rels_path = f"xl/worksheets/_rels/sheet{sheet_index}.xml.rels"
        if sheet_rels_path not in zf.namelist():
            return results

        rels_root = ET.fromstring(zf.read(sheet_rels_path))
        drawing_target: str | None = None
        for rel in rels_root:
            if rel.get("Type", "").endswith("/drawing"):
                drawing_target = rel.get("Target", "")
                break
        if not drawing_target:
            return results

        # Target may be absolute ("/xl/drawings/drawing1.xml") or relative
        # ("../drawings/drawing1.xml" from xl/worksheets/).  Normalize both to
        # the ZIP-internal path "xl/drawings/drawing1.xml".
        if drawing_target.startswith("/"):
            drawing_zip_path = drawing_target.lstrip("/")
        else:
            drawing_zip_path = "xl/" + drawing_target.lstrip("../")
        drawing_name = drawing_zip_path.split("/")[-1]
        drawing_dir = "/".join(drawing_zip_path.split("/")[:-1])

        if drawing_zip_path not in zf.namelist():
            return results

        drawing_rels_path = f"{drawing_dir}/_rels/{drawing_name}.rels"
        if drawing_rels_path not in zf.namelist():
            return results

        rid_to_media: dict[str, str] = {}
        for rel in ET.fromstring(zf.read(drawing_rels_path)):
            rid = rel.get("Id", "")
            if rid:
                rid_to_media[rid] = rel.get("Target", "")

        drawing_root = ET.fromstring(zf.read(drawing_zip_path))
        anchor_tags = (
            f"{{{_NS_XDR}}}twoCellAnchor",
            f"{{{_NS_XDR}}}oneCellAnchor",
            f"{{{_NS_XDR}}}absoluteAnchor",
        )

        for anchor in drawing_root:
            if anchor.tag not in anchor_tags:
                continue

            anchor_row = 0
            from_el = anchor.find(f"{{{_NS_XDR}}}from")
            if from_el is not None:
                row_el = from_el.find(f"{{{_NS_XDR}}}row")
                if row_el is not None and row_el.text is not None:
                    try:
                        anchor_row = int(row_el.text)
                    except ValueError:
                        pass

            # iter() searches recursively, so pics inside grpSp are found too.
            seen_rids: set[str] = set()
            for pic in anchor.iter(f"{{{_NS_XDR}}}pic"):
                blip = pic.find(f".//{{{_NS_A}}}blip")
                if blip is None:
                    continue
                embed_id = blip.get(f"{{{_NS_R}}}embed", "")
                if not embed_id or embed_id in seen_rids:
                    continue
                seen_rids.add(embed_id)

                media_path = rid_to_media.get(embed_id, "")
                if not media_path:
                    continue

                ext = media_path.rsplit(".", 1)[-1].lower() if "." in media_path else ""
                if ext not in _SUPPORTED_IMAGE_FORMATS:
                    continue

                if media_path.startswith("/"):
                    media_zip_path = media_path.lstrip("/")
                else:
                    media_zip_path = "xl/" + media_path.lstrip("../")
                if media_zip_path not in zf.namelist():
                    continue

                try:
                    img_bytes = zf.read(media_zip_path)
                except Exception:
                    continue

                asset_key = f"images/{safe_name}_image_{counter}.{ext}"
                results.append((anchor_row, asset_key, img_bytes))
                counter += 1

    return results


def _trim_cols(rows: list) -> list:
    """Trim trailing empty columns while keeping row count intact for anchor alignment."""
    if not rows:
        return []
    max_col = -1
    for row in rows:
        for i, v in enumerate(row):
            if v is not None:
                max_col = max(max_col, i)
    if max_col < 0:
        return rows  # All cells empty; canvas detection will handle it
    return [list(r)[:max_col + 1] for r in rows]


def _is_canvas(rows: list) -> bool:
    """Return True if this sheet is a drawing canvas rather than a data table.

    A canvas sheet is one where the cell grid is used as a background for
    floating shapes rather than to store structured data.  Heuristic: >80%
    of cells empty AND wider than 15 columns (so narrow sparse tables are
    not mis-classified).
    """
    if not rows:
        return True
    total = sum(len(r) for r in rows)
    if total == 0:
        return True
    filled = sum(1 for r in rows for v in r if v is not None)
    col_count = max(len(r) for r in rows)
    return (filled / total) < 0.2 and col_count > 15


def _extract_shape_texts(
    zip_bytes: bytes, sheet_index: int
) -> list[tuple[int, str]]:
    """Extract text from shape (sp) elements in the drawing XML.

    Returns (anchor_row_0based, text) for shapes that have non-empty text.
    Used for canvas sheets to produce a readable text outline of the drawn
    content instead of a massive sparse table.
    """
    results: list[tuple[int, str]] = []

    try:
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except zipfile.BadZipFile:
        return results

    with zf:
        sheet_rels_path = f"xl/worksheets/_rels/sheet{sheet_index}.xml.rels"
        if sheet_rels_path not in zf.namelist():
            return results

        rels_root = ET.fromstring(zf.read(sheet_rels_path))
        drawing_target: str | None = None
        for rel in rels_root:
            if rel.get("Type", "").endswith("/drawing"):
                drawing_target = rel.get("Target", "")
                break
        if not drawing_target:
            return results

        if drawing_target.startswith("/"):
            drawing_zip_path = drawing_target.lstrip("/")
        else:
            drawing_zip_path = "xl/" + drawing_target.lstrip("../")

        if drawing_zip_path not in zf.namelist():
            return results

        drawing_root = ET.fromstring(zf.read(drawing_zip_path))
        anchor_tags = (
            f"{{{_NS_XDR}}}twoCellAnchor",
            f"{{{_NS_XDR}}}oneCellAnchor",
            f"{{{_NS_XDR}}}absoluteAnchor",
        )

        for anchor in drawing_root:
            if anchor.tag not in anchor_tags:
                continue

            anchor_row = 0
            from_el = anchor.find(f"{{{_NS_XDR}}}from")
            if from_el is not None:
                row_el = from_el.find(f"{{{_NS_XDR}}}row")
                if row_el is not None and row_el.text is not None:
                    try:
                        anchor_row = int(row_el.text)
                    except ValueError:
                        pass

            # iter() finds sp elements recursively (handles grpSp nesting).
            for sp in anchor.iter(f"{{{_NS_XDR}}}sp"):
                txbody = sp.find(f"{{{_NS_XDR}}}txBody")
                if txbody is None:
                    continue
                lines = []
                for p in txbody.iter(f"{{{_NS_A}}}p"):
                    para = "".join(t.text or "" for t in p.iter(f"{{{_NS_A}}}t"))
                    if para.strip():
                        lines.append(para.strip())
                if lines:
                    results.append((anchor_row, "\n".join(lines)))

    return results


def _canvas_cells_to_text(rows: list) -> str:
    """Render canvas sheet cell data.

    Rows with ≥3 filled cells anchor a cluster.  The cluster grows by bridging
    up to 2 consecutive empty rows (gaps between table items) and absorbing
    single-cell continuation rows.  A cluster is rendered as a compressed
    markdown table when it meets two criteria:
      - At least MIN_TABLE_ROWS non-empty rows
      - At most MAX_DISTINCT_COLS distinct column positions across all rows
        (distinguishes compact spec tables from scattered wireframe labels)

    Column compression removes all-empty columns so a table whose data sits at
    columns 32/33/38 in the original sheet renders as a 3-column table, not a
    39-column table of mostly empty cells.

    Rows that don't join a qualifying cluster are emitted as plain text.
    """
    if not rows:
        return ""

    MIN_TABLE_ROWS = 4
    MAX_DISTINCT_COLS = 7
    MAX_GAP = 2          # consecutive empty rows allowed inside a cluster
    ANCHOR_CELLS = 3     # min filled cells for a row to start a cluster

    def _compress(cluster: list) -> list:
        max_c = max((len(r) for r in cluster), default=0)
        active = [c for c in range(max_c)
                  if any(c < len(r) and r[c] is not None and str(r[c]).strip()
                         for r in cluster)]
        if not active:
            return cluster
        return [[r[c] if c < len(r) else None for c in active] for r in cluster]

    sections: list[str] = []
    i = 0

    while i < len(rows):
        row = rows[i]
        filled = [v for v in row if v is not None and str(v).strip()]

        if len(filled) >= ANCHOR_CELLS:
            # This row anchors a cluster; collect it with gap-bridging.
            cluster = [row]
            j = i + 1
            consecutive_empty = 0

            while j < len(rows):
                nxt_filled = [v for v in rows[j] if v is not None and str(v).strip()]
                if nxt_filled:
                    cluster.append(rows[j])
                    consecutive_empty = 0
                    j += 1
                else:
                    consecutive_empty += 1
                    if consecutive_empty > MAX_GAP:
                        break
                    cluster.append(rows[j])  # placeholder kept to bridge the gap
                    j += 1

            # Drop trailing empty rows (only kept to count consecutive empties).
            while cluster and not any(
                v is not None and str(v).strip() for v in cluster[-1]
            ):
                cluster.pop()

            # Count distinct column positions that carry data anywhere in the cluster.
            all_cols: set[int] = set()
            for r in cluster:
                for ci, v in enumerate(r):
                    if v is not None and str(v).strip():
                        all_cols.add(ci)

            if len(cluster) >= MIN_TABLE_ROWS and len(all_cols) <= MAX_DISTINCT_COLS:
                md = _rows_to_markdown(_compress(cluster))
                if md:
                    sections.append(md)
            else:
                for r in cluster:
                    vals = [str(v).strip() for v in r if v is not None and str(v).strip()]
                    if vals:
                        sections.append(" | ".join(vals))
            i = j

        else:
            vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
            if vals:
                sections.append(" ".join(vals))
            i += 1

    return "\n\n".join(s for s in sections if s)


def _interleave(rows: list, insertions: list[tuple[int, str]]) -> str:
    """Interleave row-table segments with image/comment insertions at their anchor rows.

    Each insertion is (anchor_row_0based, text).  An anchor row of N means the
    insertion appears between rows[N-1] and rows[N]: the first fragment covers
    rows[0:N] and the next fragment starts at rows[N:].

    Insertions at the same anchor row are grouped and emitted together.
    Empty fragments (no rows) are skipped so we don't emit blank tables.
    """
    if not insertions:
        return _rows_to_markdown(rows)

    # Sort by anchor row so we can walk forward through rows once.
    sorted_insertions = sorted(insertions, key=lambda t: t[0])

    parts: list[str] = []
    row_cursor = 0

    # Group insertions by anchor row so duplicates don't cause double splits.
    i = 0
    while i < len(sorted_insertions):
        anchor_row = sorted_insertions[i][0]

        # Clamp to valid range.
        split_at = max(row_cursor, min(anchor_row, len(rows)))

        # Emit the segment of rows before this insertion point.
        segment = rows[row_cursor:split_at]
        if segment:
            parts.append(_rows_to_markdown(segment))
        row_cursor = split_at

        # Emit all insertions sharing this anchor row.
        while i < len(sorted_insertions) and sorted_insertions[i][0] == anchor_row:
            parts.append(sorted_insertions[i][1])
            i += 1

    # Emit any remaining rows after the last insertion.
    tail = rows[row_cursor:]
    if tail:
        parts.append(_rows_to_markdown(tail))

    return "\n\n".join(parts)


def _fill_merged_cells(ws) -> None:
    """Copy the top-left value of each merge region into every cell it covers."""
    for merge in list(ws.merged_cells.ranges):
        top_left = ws.cell(merge.min_row, merge.min_col).value
        ws.unmerge_cells(str(merge))
        for row in range(merge.min_row, merge.max_row + 1):
            for col in range(merge.min_col, merge.max_col + 1):
                ws.cell(row, col).value = top_left


def _rows_to_markdown(rows: list) -> str:
    if not rows:
        return ""

    rows = [r for r in rows if any(v is not None for v in r)]
    if not rows:
        return ""

    col_count = max(len(r) for r in rows)

    def pad(row):
        return list(row) + [None] * (col_count - len(row))

    def cell(v):
        return str(v).replace("|", "\\|").replace("\n", " ") if v is not None else ""

    header = pad(rows[0])
    lines = []
    lines.append("| " + " | ".join(cell(c) for c in header) + " |")
    lines.append("| " + " | ".join("---" for _ in header) + " |")
    for row in rows[1:]:
        lines.append("| " + " | ".join(cell(c) for c in pad(row)) + " |")

    return "\n".join(lines)
